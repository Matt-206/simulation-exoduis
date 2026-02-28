"""
Comprehensive Spreadsheet Exporter for the Japan Exodus Simulation.

Produces a multi-sheet Excel workbook (.xlsx) with full statistics across
every dimension of the simulation: demographics, geography, migration,
PCAF dynamics, agent distributions, policy outcomes, and summary stats.

Usage:
    python export_spreadsheet.py --years 20 --scale 200
    python export_spreadsheet.py --years 40 --scale 200 --scenario optimistic
    python export_spreadsheet.py --years 10 --scale 500 --snapshot output/branch/branch_snapshot.pkl
"""

import argparse
import sys
import time
from pathlib import Path
from collections import defaultdict

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.chart import LineChart, Reference, BarChart, ScatterChart
from openpyxl.chart.series import SeriesLabel
from openpyxl.utils import get_column_letter

from model.config import SimulationConfig, ScaleConfig, disable_gpu
from model.model import ExodusModel

TIER_NAMES = {0: "Tokyo", 1: "Core City", 2: "Periphery"}


# ======================================================================
# Style definitions
# ======================================================================
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUBHEADER_FONT = Font(bold=True, size=10)
TITLE_FONT = Font(bold=True, size=14, color="1F3864")
NUM_FMT_INT = "#,##0"
NUM_FMT_PCT = "0.00%"
NUM_FMT_DEC2 = "#,##0.00"
NUM_FMT_DEC4 = "0.0000"
NUM_FMT_MONEY = "#,##0"
THIN_BORDER = Border(
    bottom=Side(style="thin", color="B4C6E7"),
)


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def auto_width(ws, min_width=10, max_width=30):
    for col in ws.columns:
        max_len = min_width
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[col_letter].width = max_len


def add_chart(ws, title, min_col, max_col, min_row, max_row, cat_col, anchor, width=18, height=10, chart_type="line"):
    if chart_type == "line":
        chart = LineChart()
    elif chart_type == "bar":
        chart = BarChart()
    else:
        chart = LineChart()
    chart.title = title
    chart.style = 10
    chart.width = width
    chart.height = height
    cats = Reference(ws, min_col=cat_col, min_row=min_row + 1, max_row=max_row)
    for c in range(min_col, max_col + 1):
        vals = Reference(ws, min_col=c, min_row=min_row, max_row=max_row)
        chart.add_data(vals, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, anchor)


# ======================================================================
# Data collectors (run during simulation)
# ======================================================================
class FullDataCollector:
    """Collects deep data at annual boundaries during simulation."""

    def __init__(self, model: ExodusModel, scale: int):
        self.model = model
        self.scale = scale
        self._annual_snapshots = []
        self._location_snapshots = []
        self._age_pyramids = []
        self._tier_demographics = []
        self._prefecture_data = []
        self._migration_tier_matrix = []
        self._last_year = None

    def collect_if_annual(self):
        yr = self.model.current_year
        q = self.model.current_step_in_year
        if yr == self._last_year:
            return
        if q != 0 and self.model.total_steps > 0:
            return
        self._last_year = yr
        self._collect_annual(yr)

    def _collect_annual(self, yr):
        model = self.model
        pool = model.agents_pool
        loc = model.loc_state
        n = pool.next_id
        alive = pool.alive[:n]
        scale = self.scale

        ages = pool.age[:n][alive]
        sexes = pool.sex[:n][alive]
        incomes = pool.income[:n][alive]
        locs = pool.location[:n][alive]
        edu = pool.education[:n][alive]
        marital = pool.marital_status[:n][alive]
        children = pool.n_children[:n][alive]
        remote = pool.remote_worker[:n][alive]
        unemployed = pool.unemployed[:n][alive]
        tiers = loc.tier[locs]
        origin_tiers = pool.origin_tier[:n][alive]

        n_alive = int(alive.sum())

        # Age pyramid (5-year bins)
        for age_lo in range(0, 105, 5):
            age_hi = age_lo + 4
            mask = (ages >= age_lo) & (ages <= age_hi)
            males = int(((sexes == 0) & mask).sum()) * scale
            females = int(((sexes == 1) & mask).sum()) * scale
            self._age_pyramids.append({
                "year": yr, "age_group": f"{age_lo}-{age_hi}",
                "males": males, "females": females, "total": males + females,
            })

        # Tier demographics
        for tier_id, tier_name in TIER_NAMES.items():
            tmask = tiers == tier_id
            if tmask.sum() == 0:
                continue
            t_ages = ages[tmask]
            t_inc = incomes[tmask]
            t_mar = marital[tmask]
            t_edu = edu[tmask]
            t_child = children[tmask]
            t_remote = remote[tmask]
            t_unemp = unemployed[tmask]
            t_pop = int(tmask.sum()) * scale

            married_count = int((t_mar == 1).sum())
            edu_high = int((t_edu >= 2).sum())
            remote_count = int(t_remote.sum())
            unemp_count = int(t_unemp.sum())
            young = int(((t_ages >= 20) & (t_ages <= 34)).sum())
            elderly = int((t_ages >= 65).sum())

            # Retention: agents born in this tier still here
            origin_match = (origin_tiers[tmask] == tier_id)
            retained = int(origin_match.sum()) * scale

            self._tier_demographics.append({
                "year": yr, "tier": tier_name,
                "population": t_pop,
                "mean_age": round(float(t_ages.mean()), 1),
                "median_age": round(float(np.median(t_ages)), 1),
                "pct_young_20_34": round(young / max(tmask.sum(), 1), 4),
                "pct_elderly_65plus": round(elderly / max(tmask.sum(), 1), 4),
                "dependency_ratio": round(elderly / max(young, 1), 3),
                "mean_income_jpy": round(float(t_inc.mean())),
                "median_income_jpy": round(float(np.median(t_inc))),
                "pct_married": round(married_count / max(tmask.sum(), 1), 4),
                "mean_children": round(float(t_child.mean()), 2),
                "pct_higher_education": round(edu_high / max(tmask.sum(), 1), 4),
                "pct_remote_work": round(remote_count / max(tmask.sum(), 1), 4),
                "pct_unemployed": round(unemp_count / max(tmask.sum(), 1), 4),
                "retained_from_origin": retained,
            })

        # Per-location snapshot
        names = getattr(loc, '_municipality_names', None)
        prefs = getattr(loc, '_prefectures', None)
        n_locs = len(loc.tier)
        for i in range(n_locs):
            row = {
                "year": yr,
                "location_id": i,
                "name": names[i] if names else f"Loc_{i}",
                "prefecture": str(prefs[i]) if prefs is not None else "N/A",
                "tier": TIER_NAMES.get(int(loc.tier[i]), "?"),
                "population": int(loc.population[i]) * scale,
                "capacity": int(loc.capacity[i]) * scale,
                "occupancy_rate": round(loc.population[i] / max(loc.capacity[i], 1), 4),
                "prestige": round(float(loc.prestige[i]), 4),
                "convenience": round(float(loc.convenience[i]), 4),
                "anomie": round(float(loc.anomie[i]), 4),
                "financial_friction": round(float(loc.financial_friction[i]), 4),
                "rent_index": round(float(loc.rent_index[i]), 4),
                "vacancy_rate": round(float(loc.vacancy_rate[i]), 4),
                "healthcare": round(float(loc.healthcare_score[i]), 4),
                "childcare": round(float(loc.childcare_score[i]), 4),
                "digital_access": round(float(loc.digital_access[i]), 4),
                "hq_count": int(loc.hq_count[i]),
                "school_closed": bool(loc.school_closed[i]) if loc.school_closed is not None else False,
                "lon": round(float(loc.lon[i]), 4),
                "lat": round(float(loc.lat[i]), 4),
            }
            if loc.depopulation_rate is not None:
                row["depop_rate_quarterly"] = round(float(loc.depopulation_rate[i]), 6)
            if loc.social_friction is not None:
                row["social_friction"] = round(float(loc.social_friction[i]), 4)
            self._location_snapshots.append(row)

        # Prefecture aggregation
        if prefs is not None:
            for pref in np.unique(prefs):
                pmask = prefs == pref
                p_pop = int(loc.population[pmask].sum()) * scale
                p_locs = int(pmask.sum())
                p_tier_counts = {
                    "n_tokyo": int((loc.tier[pmask] == 0).sum()),
                    "n_core": int((loc.tier[pmask] == 1).sum()),
                    "n_periphery": int((loc.tier[pmask] == 2).sum()),
                }
                p_hq = int(loc.hq_count[pmask].sum())
                p_schools_closed = int(loc.school_closed[pmask].sum()) if loc.school_closed is not None else 0
                self._prefecture_data.append({
                    "year": yr,
                    "prefecture": str(pref),
                    "population": p_pop,
                    "n_municipalities": p_locs,
                    **p_tier_counts,
                    "mean_prestige": round(float(loc.prestige[pmask].mean()), 4),
                    "mean_convenience": round(float(loc.convenience[pmask].mean()), 4),
                    "mean_anomie": round(float(loc.anomie[pmask].mean()), 4),
                    "mean_financial_friction": round(float(loc.financial_friction[pmask].mean()), 4),
                    "total_hq": p_hq,
                    "schools_closed": p_schools_closed,
                    "mean_vacancy": round(float(loc.vacancy_rate[pmask].mean()), 4),
                    "mean_healthcare": round(float(loc.healthcare_score[pmask].mean()), 4),
                    "mean_digital_access": round(float(loc.digital_access[pmask].mean()), 4),
                })

        # Migration tier-to-tier matrix this year
        ct = model.cannibalism_tracker
        self._migration_tier_matrix.append({
            "year": yr,
            "periphery_to_core": ct["periphery_to_core"],
            "periphery_to_tokyo": ct["periphery_to_tokyo"],
            "core_to_tokyo": ct["core_to_tokyo"],
            "core_to_tokyo_from_periphery": ct["core_to_tokyo_from_periphery"],
            "launchpad_ratio": ct["core_to_tokyo_from_periphery"] / max(ct["core_to_tokyo"], 1),
        })


# ======================================================================
# Excel workbook builder
# ======================================================================
def build_workbook(
    history_df: pd.DataFrame,
    flows_df: pd.DataFrame,
    collector: FullDataCollector,
    model: ExodusModel,
    scale: int,
    scenario: str,
    runtime_s: float,
) -> Workbook:
    wb = Workbook()

    # -- Sheet 1: Executive Summary --
    _build_summary_sheet(wb, history_df, model, scale, scenario, runtime_s)
    # -- Sheet 2: Annual Time Series --
    _build_timeseries_sheet(wb, history_df, scale)
    # -- Sheet 3: Quarterly Detail --
    _build_quarterly_sheet(wb, history_df, scale)
    # -- Sheet 4: Tier Demographics --
    _build_tier_sheet(wb, collector)
    # -- Sheet 5: Age Pyramids --
    _build_pyramid_sheet(wb, collector)
    # -- Sheet 6: Prefecture Dashboard --
    _build_prefecture_sheet(wb, collector)
    # -- Sheet 7: Municipality Detail (final year) --
    _build_municipality_sheet(wb, collector)
    # -- Sheet 8: PCAF Dynamics --
    _build_pcaf_sheet(wb, history_df)
    # -- Sheet 9: Migration Analysis --
    _build_migration_sheet(wb, history_df, flows_df, collector)
    # -- Sheet 10: Policy Metrics --
    _build_policy_sheet(wb, history_df)
    # -- Sheet 11: Agent Distribution Snapshot --
    _build_agent_distribution_sheet(wb, model, scale)
    # -- Sheet 12: Human Warehouse Index --
    _build_warehouse_sheet(wb, collector)

    return wb


def _build_summary_sheet(wb, history_df, model, scale, scenario, runtime_s):
    ws = wb.active
    ws.title = "Executive Summary"
    ws.sheet_properties.tabColor = "1F3864"

    ws.merge_cells("A1:F1")
    ws["A1"] = "JAPAN EXODUS SIMULATION -- RESEARCH DATA EXPORT"
    ws["A1"].font = TITLE_FONT

    ws.merge_cells("A2:F2")
    ws["A2"] = f"Scenario: {scenario.upper()} | Scale: 1:{scale} | Generated: {time.strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(italic=True, size=10, color="666666")

    row = 4
    summary_data = []

    h = history_df
    annual = h.groupby("year").last().reset_index()
    first = annual.iloc[0] if len(annual) else {}
    last = annual.iloc[-1] if len(annual) else {}

    start_pop = int(first.get("total_population", 0)) * scale
    end_pop = int(last.get("total_population", 0)) * scale
    pop_change = end_pop - start_pop
    years_sim = int(last.get("year", 0)) - int(first.get("year", 0))
    total_births = int(h["births"].sum()) * scale
    total_deaths = int(h["deaths"].sum()) * scale
    total_migrations = int(h["migrations"].sum())

    summary_data = [
        ("POPULATION", ""),
        ("Start Year", int(first.get("year", 0))),
        ("End Year", int(last.get("year", 0))),
        ("Simulation Duration", f"{years_sim} years"),
        ("Starting Population", f"{start_pop:,}"),
        ("Ending Population", f"{end_pop:,}"),
        ("Population Change", f"{pop_change:+,} ({pop_change/max(start_pop,1)*100:+.1f}%)"),
        ("Total Births", f"{total_births:,}"),
        ("Total Deaths", f"{total_deaths:,}"),
        ("Natural Increase", f"{total_births - total_deaths:+,}"),
        ("Total Migrations", f"{total_migrations:,}"),
        ("", ""),
        ("SPATIAL DISTRIBUTION (Final Year)", ""),
        ("Tokyo Population Share", f"{float(last.get('tokyo_pop_share', 0))*100:.2f}%"),
        ("Core City Share", f"{float(last.get('core_pop_share', 0))*100:.2f}%"),
        ("Periphery Share", f"{float(last.get('peri_pop_share', 0))*100:.2f}%"),
        ("Tokyo Pref Share", f"{float(last.get('tokyo_pref_share', 0))*100:.2f}%"),
        ("", ""),
        ("FERTILITY & MORTALITY", ""),
        ("Final TFR Proxy", f"{float(last.get('tfr_proxy', 0)):.3f}"),
        ("Final Mean Age", f"{float(last.get('mean_age', 0)):.1f}"),
        ("% Married (final)", f"{float(last.get('pct_married', 0))*100:.1f}%"),
        ("", ""),
        ("MIGRATION DYNAMICS (Cumulative)", ""),
        ("Periphery -> Core", f"{int(last.get('periphery_to_core_flow', 0)):,}"),
        ("Periphery -> Tokyo", f"{int(last.get('periphery_to_tokyo_flow', 0)):,}"),
        ("Core -> Tokyo", f"{int(last.get('core_to_tokyo_flow', 0)):,}"),
        ("Launchpad Ratio", f"{float(last.get('launchpad_ratio', 0)):.3f}"),
        ("Cannibalism Ratio", f"{float(last.get('cannibalism_ratio', 0)):.3f}"),
        ("", ""),
        ("STRUCTURAL INDICATORS (Final Year)", ""),
        ("School Closures (periphery)", int(last.get("n_school_closures", 0))),
        ("Human Warehouse Towns", int(last.get("n_human_warehouse_towns", 0))),
        ("Unemployed Agents", int(last.get("n_unemployed", 0)) * scale),
        ("HQs in Tokyo", int(last.get("tokyo_hq_count", 0))),
        ("HQs in Core Cities", int(last.get("core_hq_count", 0))),
        ("", ""),
        ("SIMULATION METADATA", ""),
        ("Agent Scale", f"1:{scale}"),
        ("Agent Count", f"{model.agents_pool.n_alive:,}"),
        ("Municipality Count", len(model.loc_state.tier)),
        ("Runtime", f"{runtime_s:.1f} seconds"),
    ]

    for label, value in summary_data:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        if label and not value:
            ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="2F5496")
            ws.cell(row=row, column=1).fill = SUBHEADER_FILL
            ws.cell(row=row, column=2).fill = SUBHEADER_FILL
        row += 1

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 25


def _build_timeseries_sheet(wb, history_df, scale):
    ws = wb.create_sheet("Annual Time Series")
    ws.sheet_properties.tabColor = "2E75B6"

    annual = history_df.groupby("year").agg({
        "total_population": "last",
        "tokyo_population": "last",
        "core_population": "last",
        "periphery_population": "last",
        "births": "sum",
        "deaths": "sum",
        "marriages": "sum",
        "migrations": "sum",
        "tfr_proxy": "last",
        "mean_age": "last",
        "pct_married": "last",
        "mean_income": "last",
        "pct_remote": "last",
        "tokyo_pop_share": "last",
        "core_pop_share": "last",
        "peri_pop_share": "last",
        "periphery_to_core_flow": "last",
        "periphery_to_tokyo_flow": "last",
        "core_to_tokyo_flow": "last",
        "cannibalism_ratio": "last",
        "launchpad_ratio": "last",
        "n_school_closures": "last",
        "n_human_warehouse_towns": "last",
        "n_unemployed": "last",
        "mean_prestige_tokyo": "last",
        "mean_prestige_core": "last",
        "mean_prestige_periphery": "last",
        "mean_convenience_periphery": "last",
        "mean_anomie_periphery": "last",
        "mean_friction_tokyo": "last",
    }).reset_index()

    # Scale population columns
    pop_cols = ["total_population", "tokyo_population", "core_population",
                "periphery_population", "births", "deaths", "n_unemployed"]
    for c in pop_cols:
        if c in annual.columns:
            annual[c] = annual[c] * scale

    # Natural increase
    annual["natural_increase"] = annual["births"] - annual["deaths"]
    annual["growth_rate_pct"] = annual["total_population"].pct_change() * 100

    cols = list(annual.columns)
    headers = [c.replace("_", " ").title() for c in cols]

    for j, h in enumerate(headers, 1):
        ws.cell(row=1, column=j, value=h)
    style_header_row(ws, 1, len(headers))

    for i, row_data in annual.iterrows():
        for j, col in enumerate(cols, 1):
            cell = ws.cell(row=i + 2, column=j, value=row_data[col])
            if "population" in col or col in ("births", "deaths", "natural_increase", "n_unemployed"):
                cell.number_format = NUM_FMT_INT
            elif "share" in col or "pct" in col or "rate" in col:
                cell.number_format = NUM_FMT_PCT if "share" in col else NUM_FMT_DEC4
            elif "ratio" in col:
                cell.number_format = NUM_FMT_DEC4
            elif "income" in col:
                cell.number_format = NUM_FMT_MONEY

    auto_width(ws)

    # Population chart
    n_rows = len(annual) + 1
    if n_rows > 2:
        add_chart(ws, "Population Trajectory", 2, 2, 1, n_rows, 1, f"A{n_rows + 3}")
        add_chart(ws, "Births & Deaths", 6, 7, 1, n_rows, 1, f"A{n_rows + 20}")


def _build_quarterly_sheet(wb, history_df, scale):
    ws = wb.create_sheet("Quarterly Detail")
    ws.sheet_properties.tabColor = "548235"

    df = history_df.copy()
    pop_cols = ["total_population", "tokyo_population", "core_population",
                "periphery_population", "births", "deaths", "n_unemployed"]
    for c in pop_cols:
        if c in df.columns:
            df[c] = df[c] * scale

    cols = list(df.columns)
    headers = [c.replace("_", " ").title() for c in cols]

    for j, h in enumerate(headers, 1):
        ws.cell(row=1, column=j, value=h)
    style_header_row(ws, 1, len(headers))

    for i, (_, row_data) in enumerate(df.iterrows()):
        for j, col in enumerate(cols, 1):
            cell = ws.cell(row=i + 2, column=j, value=row_data[col])
            if "population" in col or col in ("births", "deaths", "n_unemployed", "marriages", "migrations"):
                cell.number_format = NUM_FMT_INT

    auto_width(ws)


def _build_tier_sheet(wb, collector):
    ws = wb.create_sheet("Tier Demographics")
    ws.sheet_properties.tabColor = "C00000"

    df = pd.DataFrame(collector._tier_demographics)
    if df.empty:
        ws["A1"] = "No tier data collected"
        return

    cols = list(df.columns)
    headers = [c.replace("_", " ").title() for c in cols]

    for j, h in enumerate(headers, 1):
        ws.cell(row=1, column=j, value=h)
    style_header_row(ws, 1, len(headers))

    for i, (_, row_data) in enumerate(df.iterrows()):
        for j, col in enumerate(cols, 1):
            cell = ws.cell(row=i + 2, column=j, value=row_data[col])
            if "population" in col or "retained" in col:
                cell.number_format = NUM_FMT_INT
            elif "pct" in col or "ratio" in col:
                cell.number_format = NUM_FMT_PCT
            elif "income" in col:
                cell.number_format = NUM_FMT_MONEY

    auto_width(ws)


def _build_pyramid_sheet(wb, collector):
    ws = wb.create_sheet("Age Pyramids")
    ws.sheet_properties.tabColor = "7030A0"

    df = pd.DataFrame(collector._age_pyramids)
    if df.empty:
        ws["A1"] = "No age pyramid data"
        return

    cols = list(df.columns)
    headers = [c.replace("_", " ").title() for c in cols]

    for j, h in enumerate(headers, 1):
        ws.cell(row=1, column=j, value=h)
    style_header_row(ws, 1, len(headers))

    for i, (_, row_data) in enumerate(df.iterrows()):
        for j, col in enumerate(cols, 1):
            cell = ws.cell(row=i + 2, column=j, value=row_data[col])
            if col in ("males", "females", "total"):
                cell.number_format = NUM_FMT_INT

    auto_width(ws)


def _build_prefecture_sheet(wb, collector):
    ws = wb.create_sheet("Prefecture Dashboard")
    ws.sheet_properties.tabColor = "ED7D31"

    df = pd.DataFrame(collector._prefecture_data)
    if df.empty:
        ws["A1"] = "No prefecture data"
        return

    cols = list(df.columns)
    headers = [c.replace("_", " ").title() for c in cols]

    for j, h in enumerate(headers, 1):
        ws.cell(row=1, column=j, value=h)
    style_header_row(ws, 1, len(headers))

    for i, (_, row_data) in enumerate(df.iterrows()):
        for j, col in enumerate(cols, 1):
            cell = ws.cell(row=i + 2, column=j, value=row_data[col])
            if "population" in col or "hq" in col.lower() or "schools" in col:
                cell.number_format = NUM_FMT_INT
            elif "mean" in col:
                cell.number_format = NUM_FMT_DEC4

    auto_width(ws)


def _build_municipality_sheet(wb, collector):
    ws = wb.create_sheet("Municipality Detail")
    ws.sheet_properties.tabColor = "70AD47"

    all_locs = pd.DataFrame(collector._location_snapshots)
    if all_locs.empty:
        ws["A1"] = "No location data"
        return

    last_year = all_locs["year"].max()
    df = all_locs[all_locs["year"] == last_year].copy()
    df = df.sort_values("population", ascending=False)

    cols = list(df.columns)
    headers = [c.replace("_", " ").title() for c in cols]

    for j, h in enumerate(headers, 1):
        ws.cell(row=1, column=j, value=h)
    style_header_row(ws, 1, len(headers))

    for i, (_, row_data) in enumerate(df.iterrows()):
        for j, col in enumerate(cols, 1):
            val = row_data[col]
            cell = ws.cell(row=i + 2, column=j, value=val)
            if col in ("population", "capacity", "hq_count"):
                cell.number_format = NUM_FMT_INT
            elif col in ("prestige", "convenience", "anomie", "financial_friction",
                         "vacancy_rate", "healthcare", "childcare", "digital_access",
                         "occupancy_rate", "rent_index", "social_friction"):
                cell.number_format = NUM_FMT_DEC4

    auto_width(ws)

    # Also add a sheet with ALL years
    ws2 = wb.create_sheet("Municipality All Years")
    ws2.sheet_properties.tabColor = "A9D18E"
    df2 = all_locs.sort_values(["year", "location_id"])
    cols2 = list(df2.columns)
    for j, h in enumerate([c.replace("_", " ").title() for c in cols2], 1):
        ws2.cell(row=1, column=j, value=h)
    style_header_row(ws2, 1, len(cols2))
    for i, (_, row_data) in enumerate(df2.iterrows()):
        for j, col in enumerate(cols2, 1):
            ws2.cell(row=i + 2, column=j, value=row_data[col])
    auto_width(ws2)


def _build_pcaf_sheet(wb, history_df):
    ws = wb.create_sheet("PCAF Dynamics")
    ws.sheet_properties.tabColor = "BF8F00"

    annual = history_df.groupby("year").last().reset_index()
    pcaf_cols = [
        "year", "mean_prestige_tokyo", "mean_prestige_core", "mean_prestige_periphery",
        "mean_convenience_periphery", "mean_anomie_periphery", "mean_friction_tokyo",
        "n_human_warehouse_towns", "insurance_surcharge_jpy",
    ]
    available = [c for c in pcaf_cols if c in annual.columns]
    df = annual[available]

    cols = list(df.columns)
    headers = [c.replace("_", " ").title() for c in cols]

    for j, h in enumerate(headers, 1):
        ws.cell(row=1, column=j, value=h)
    style_header_row(ws, 1, len(headers))

    for i, (_, row_data) in enumerate(df.iterrows()):
        for j, col in enumerate(cols, 1):
            cell = ws.cell(row=i + 2, column=j, value=row_data[col])
            if "mean" in col:
                cell.number_format = NUM_FMT_DEC4
            elif "surcharge" in col:
                cell.number_format = NUM_FMT_MONEY

    auto_width(ws)

    n_rows = len(df) + 1
    if n_rows > 2:
        add_chart(ws, "Prestige by Tier", 2, 4, 1, n_rows, 1, f"A{n_rows + 3}")


def _build_migration_sheet(wb, history_df, flows_df, collector):
    ws = wb.create_sheet("Migration Analysis")
    ws.sheet_properties.tabColor = "4472C4"

    # Tier-to-tier cumulative
    annual = history_df.groupby("year").last().reset_index()
    mig_cols = ["year", "periphery_to_core_flow", "periphery_to_tokyo_flow",
                "core_to_tokyo_flow", "core_to_tokyo_from_periphery",
                "cannibalism_ratio", "launchpad_ratio", "migrations"]
    available = [c for c in mig_cols if c in annual.columns]
    df = annual[available]

    cols = list(df.columns)
    headers = [c.replace("_", " ").title() for c in cols]
    for j, h in enumerate(headers, 1):
        ws.cell(row=1, column=j, value=h)
    style_header_row(ws, 1, len(headers))

    for i, (_, row_data) in enumerate(df.iterrows()):
        for j, col in enumerate(cols, 1):
            cell = ws.cell(row=i + 2, column=j, value=row_data[col])
            if "ratio" in col:
                cell.number_format = NUM_FMT_DEC4

    auto_width(ws)

    # Full migration flow log on separate sheet (if small enough)
    if not flows_df.empty and len(flows_df) <= 500_000:
        ws2 = wb.create_sheet("Migration Flows (Raw)")
        ws2.sheet_properties.tabColor = "8DB4E2"
        cols2 = list(flows_df.columns)
        for j, h in enumerate([c.replace("_", " ").title() for c in cols2], 1):
            ws2.cell(row=1, column=j, value=h)
        style_header_row(ws2, 1, len(cols2))
        for i, (_, row_data) in enumerate(flows_df.iterrows()):
            if i >= 500_000:
                break
            for j, col in enumerate(cols2, 1):
                ws2.cell(row=i + 2, column=j, value=row_data[col])
        auto_width(ws2)
    elif not flows_df.empty:
        # Too many rows -- export summary
        ws2 = wb.create_sheet("Migration Flows Summary")
        ws2.sheet_properties.tabColor = "8DB4E2"
        tier_map = {0: "Tokyo", 1: "Core", 2: "Periphery"}
        if "from_tier" in flows_df.columns and "to_tier" in flows_df.columns:
            summary = flows_df.groupby(["from_tier", "to_tier"]).size().reset_index(name="count")
            summary["from_tier_name"] = summary["from_tier"].map(tier_map)
            summary["to_tier_name"] = summary["to_tier"].map(tier_map)
            cols2 = list(summary.columns)
            for j, h in enumerate(cols2, 1):
                ws2.cell(row=1, column=j, value=h)
            style_header_row(ws2, 1, len(cols2))
            for i, (_, row_data) in enumerate(summary.iterrows()):
                for j, col in enumerate(cols2, 1):
                    ws2.cell(row=i + 2, column=j, value=row_data[col])
            auto_width(ws2)


def _build_policy_sheet(wb, history_df):
    ws = wb.create_sheet("Policy Metrics")
    ws.sheet_properties.tabColor = "FF6600"

    policy_cols = [c for c in history_df.columns if c.startswith("policy_") or c in
                   ("n_school_closures", "n_fertility_suppressed", "mean_career_anxiety_years",
                    "n_human_warehouse_towns", "in_recession", "n_disaster_locations")]
    if not policy_cols:
        ws["A1"] = "No policy-specific columns"
        return

    annual = history_df.groupby("year").last().reset_index()
    cols = ["year"] + [c for c in policy_cols if c in annual.columns]
    df = annual[cols]

    headers = [c.replace("_", " ").title() for c in df.columns]
    for j, h in enumerate(headers, 1):
        ws.cell(row=1, column=j, value=h)
    style_header_row(ws, 1, len(headers))

    for i, (_, row_data) in enumerate(df.iterrows()):
        for j, col in enumerate(list(df.columns), 1):
            ws.cell(row=i + 2, column=j, value=row_data[col])

    auto_width(ws)


def _build_agent_distribution_sheet(wb, model, scale):
    ws = wb.create_sheet("Agent Distributions")
    ws.sheet_properties.tabColor = "9DC3E6"

    pool = model.agents_pool
    loc = model.loc_state
    n = pool.next_id
    alive = pool.alive[:n]
    ages = pool.age[:n][alive]
    sexes = pool.sex[:n][alive]
    incomes = pool.income[:n][alive]
    edu = pool.education[:n][alive]
    marital = pool.marital_status[:n][alive]
    children = pool.n_children[:n][alive]
    tiers = loc.tier[pool.location[:n][alive]]
    origin = pool.origin_tier[:n][alive]

    # Income distribution by tier
    row = 1
    ws.cell(row=row, column=1, value="INCOME DISTRIBUTION BY TIER (Final Year)")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="2F5496")
    row += 1
    inc_headers = ["Tier", "Count", "Mean", "Median", "P10", "P25", "P75", "P90", "Std Dev"]
    for j, h in enumerate(inc_headers, 1):
        ws.cell(row=row, column=j, value=h)
    style_header_row(ws, row, len(inc_headers))
    row += 1

    for tier_id, tier_name in TIER_NAMES.items():
        tmask = tiers == tier_id
        if tmask.sum() == 0:
            continue
        t_inc = incomes[tmask]
        for j, val in enumerate([
            tier_name, int(tmask.sum()) * scale,
            round(float(t_inc.mean())), round(float(np.median(t_inc))),
            round(float(np.percentile(t_inc, 10))), round(float(np.percentile(t_inc, 25))),
            round(float(np.percentile(t_inc, 75))), round(float(np.percentile(t_inc, 90))),
            round(float(t_inc.std())),
        ], 1):
            cell = ws.cell(row=row, column=j, value=val)
            if j >= 2:
                cell.number_format = NUM_FMT_INT
        row += 1

    # Education distribution
    row += 2
    ws.cell(row=row, column=1, value="EDUCATION DISTRIBUTION BY TIER")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="2F5496")
    row += 1
    edu_labels = {0: "No Higher Ed", 1: "Vocational", 2: "University", 3: "Graduate"}
    edu_headers = ["Tier"] + list(edu_labels.values()) + ["% Higher Ed"]
    for j, h in enumerate(edu_headers, 1):
        ws.cell(row=row, column=j, value=h)
    style_header_row(ws, row, len(edu_headers))
    row += 1

    for tier_id, tier_name in TIER_NAMES.items():
        tmask = tiers == tier_id
        if tmask.sum() == 0:
            continue
        t_edu = edu[tmask]
        vals = [tier_name]
        total = len(t_edu)
        for e_id in sorted(edu_labels.keys()):
            vals.append(round(int((t_edu == e_id).sum()) / max(total, 1), 4))
        vals.append(round(int((t_edu >= 2).sum()) / max(total, 1), 4))
        for j, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=j, value=val)
            if j >= 2:
                cell.number_format = NUM_FMT_PCT
        row += 1

    # Marital status
    row += 2
    ws.cell(row=row, column=1, value="MARITAL STATUS BY TIER (Age 20+)")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="2F5496")
    row += 1
    mar_labels = {0: "Single", 1: "Married", 2: "Divorced", 3: "Widowed"}
    mar_headers = ["Tier"] + list(mar_labels.values())
    for j, h in enumerate(mar_headers, 1):
        ws.cell(row=row, column=j, value=h)
    style_header_row(ws, row, len(mar_headers))
    row += 1

    adult_mask = ages >= 20
    for tier_id, tier_name in TIER_NAMES.items():
        tmask = (tiers == tier_id) & adult_mask
        if tmask.sum() == 0:
            continue
        t_mar = marital[tmask]
        total = len(t_mar)
        vals = [tier_name]
        for m_id in sorted(mar_labels.keys()):
            vals.append(round(int((t_mar == m_id).sum()) / max(total, 1), 4))
        for j, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=j, value=val)
            if j >= 2:
                cell.number_format = NUM_FMT_PCT
        row += 1

    # Origin tier vs current tier (mobility matrix)
    row += 2
    ws.cell(row=row, column=1, value="ORIGIN TIER vs CURRENT TIER (Mobility Matrix)")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="2F5496")
    row += 1
    mob_headers = ["Origin \\ Current", "Tokyo", "Core City", "Periphery", "Total", "% Stayed"]
    for j, h in enumerate(mob_headers, 1):
        ws.cell(row=row, column=j, value=h)
    style_header_row(ws, row, len(mob_headers))
    row += 1

    for o_tier, o_name in TIER_NAMES.items():
        omask = origin == o_tier
        if omask.sum() == 0:
            continue
        o_tiers = tiers[omask]
        total = int(omask.sum()) * scale
        vals = [o_name]
        stayed = 0
        for c_tier in [0, 1, 2]:
            count = int((o_tiers == c_tier).sum()) * scale
            vals.append(count)
            if c_tier == o_tier:
                stayed = count
        vals.append(total)
        vals.append(round(stayed / max(total, 1), 4))
        for j, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=j, value=val)
            if j in (2, 3, 4, 5):
                cell.number_format = NUM_FMT_INT
            elif j == 6:
                cell.number_format = NUM_FMT_PCT
        row += 1

    auto_width(ws)


def _build_warehouse_sheet(wb, collector):
    ws = wb.create_sheet("Human Warehouse Index")
    ws.sheet_properties.tabColor = "843C0C"

    all_locs = pd.DataFrame(collector._location_snapshots)
    if all_locs.empty:
        ws["A1"] = "No location data"
        return

    last_year = all_locs["year"].max()
    df = all_locs[all_locs["year"] == last_year].copy()

    # High-C, High-A periphery towns
    warehouse = df[(df["tier"] == "Periphery") &
                   (df["convenience"] > 0.5) &
                   (df["anomie"] > 0.5)].sort_values("anomie", ascending=False)

    ws.cell(row=1, column=1, value="HUMAN WAREHOUSE INDEX: Periphery Towns with High Convenience + High Anomie")
    ws.cell(row=1, column=1).font = Font(bold=True, size=12, color="843C0C")

    if warehouse.empty:
        ws.cell(row=3, column=1, value="No towns meet the threshold (C > 0.5 and A > 0.5)")
        return

    show_cols = ["name", "prefecture", "population", "convenience", "anomie",
                 "prestige", "digital_access", "healthcare", "vacancy_rate",
                 "school_closed", "hq_count"]
    available = [c for c in show_cols if c in warehouse.columns]
    headers = [c.replace("_", " ").title() for c in available]

    row = 3
    for j, h in enumerate(headers, 1):
        ws.cell(row=row, column=j, value=h)
    style_header_row(ws, row, len(headers))
    row += 1

    for _, row_data in warehouse.head(200).iterrows():
        for j, col in enumerate(available, 1):
            cell = ws.cell(row=row, column=j, value=row_data[col])
            if col == "population":
                cell.number_format = NUM_FMT_INT
            elif col in ("convenience", "anomie", "prestige", "digital_access",
                         "healthcare", "vacancy_rate"):
                cell.number_format = NUM_FMT_DEC4
        row += 1

    ws.cell(row=row + 1, column=1, value=f"Total qualifying towns: {len(warehouse)}")
    ws.cell(row=row + 1, column=1).font = Font(bold=True)

    auto_width(ws)


# ======================================================================
# Main runner
# ======================================================================
def run_and_export(
    scenario: str = "baseline",
    years: int = 20,
    scale: int = 200,
    seed: int = 42,
    output_dir: str = "output",
    snapshot_path: str = None,
):
    from run import SCENARIOS

    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)

    builder = SCENARIOS.get(scenario)
    if not builder:
        print(f"Unknown scenario: {scenario}")
        sys.exit(1)

    cfg = builder(scale, years, seed)
    model = ExodusModel(cfg)

    if snapshot_path:
        model.load_snapshot(snapshot_path)
        print(f"  Loaded snapshot: {snapshot_path}")

    collector = FullDataCollector(model, scale)
    collector.collect_if_annual()

    total_steps = years * cfg.scale.steps_per_year
    print(f"\n  Running {scenario} for {years} years (scale 1:{scale})...")

    t0 = time.time()
    for step in range(total_steps):
        model.step()
        collector.collect_if_annual()

        if step % cfg.scale.steps_per_year == 0:
            yr = model.current_year
            pop = model.agents_pool.n_alive * scale
            print(f"    {yr} -- Pop: {pop:,}")

    # Final annual collection
    collector.collect_if_annual()
    runtime = time.time() - t0

    results = model.get_results()
    history_df = results["history"]
    flows_df = results["migration_flows"]

    print(f"\n  Building spreadsheet...")
    wb = build_workbook(history_df, flows_df, collector, model, scale, scenario, runtime)

    xlsx_path = str(out / f"exodus_{scenario}_{years}yr.xlsx")
    wb.save(xlsx_path)
    print(f"  Saved: {xlsx_path}")

    # Also export CSVs
    csv_dir = out / "csv"
    csv_dir.mkdir(exist_ok=True)
    history_df.to_csv(str(csv_dir / "history.csv"), index=False)
    if not flows_df.empty:
        flows_df.to_csv(str(csv_dir / "migration_flows.csv"), index=False)
    pd.DataFrame(collector._tier_demographics).to_csv(str(csv_dir / "tier_demographics.csv"), index=False)
    pd.DataFrame(collector._age_pyramids).to_csv(str(csv_dir / "age_pyramids.csv"), index=False)
    pd.DataFrame(collector._prefecture_data).to_csv(str(csv_dir / "prefecture_data.csv"), index=False)
    pd.DataFrame(collector._location_snapshots).to_csv(str(csv_dir / "municipality_snapshots.csv"), index=False)
    print(f"  CSVs saved to: {csv_dir}/")

    return wb, results


def main():
    parser = argparse.ArgumentParser(
        description="Export comprehensive simulation data to Excel spreadsheet",
    )
    parser.add_argument("--scenario", default="baseline", help="Scenario name")
    parser.add_argument("--years", type=int, default=20, help="Years to simulate")
    parser.add_argument("--scale", type=int, default=200, help="Agent scale (1:N)")
    parser.add_argument("--seed", type=int, default=42, help="Random seed")
    parser.add_argument("--output", default="output", help="Output directory")
    parser.add_argument("--snapshot", default=None, help="Load from snapshot before running")
    parser.add_argument("--no-gpu", action="store_true", help="Disable GPU")
    args = parser.parse_args()

    if args.no_gpu:
        disable_gpu()

    print("Japan Exodus -- Full Data Export")
    run_and_export(args.scenario, args.years, args.scale, args.seed, args.output, args.snapshot)


if __name__ == "__main__":
    main()
